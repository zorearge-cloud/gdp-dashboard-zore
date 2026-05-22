import streamlit as st
import pandas as pd
import requests
import io
import openpyxl
import datetime
import json
import re

# --- AYARLAR VE ANAYASA (TAM KAPSAMLI YAPI) ---
st.set_page_config(layout="wide", page_title="ZORE Veri Paneli")

# 1. KURAL: Veri çekme bağlantıları ve tab yapıları tamamen korundu
LINKS = [
    "https://docs.google.com/spreadsheets/d/1j819WkX93CkCy3VgZkSff5C_zNX5Z98jfK-FwI4ZWUU/export?format=xlsx",
    "https://docs.google.com/spreadsheets/d/1hVk6VgMFXWAukoQwMDIoOLrG8SD4UDLFFRH9VmDhXSE/export?format=xlsx",
    "https://docs.google.com/spreadsheets/d/1S1kTptWUEf705cBLw9P9mL6rrqbVjbcp1xk_hgQ-Ny0/export?format=xlsx",
    "https://docs.google.com/spreadsheets/d/1VKb6za4Fse5XrGawPG6qvrQZuFhDRGaAysmADGIC7Wc/export?format=xlsx",
    "https://docs.google.com/spreadsheets/d/1F71jiUwqvxddv7jwJibisWVaFxQ3oLQPP3CohzK_idk/export?format=xlsx"
]

TARGET_TABS = ["has_air", "has_sea", "meh_air", "meh_sea", "ist_air", "ist_sea"]

EXPECTED_COLUMNS = ['SIPARIS_TARIHI', 'FIRMA', 'TUR', 'BARKOD', 'MALIN CINSI', 'ADET', 'FIYAT', 'YUKLEME_TARIHI', 'NAKLİYE_TÜRÜ']

HEADER_MAP = {
    'SIPARIS TARIHI': 'SIPARIS_TARIHI', 'SIPARIS_TARIHI': 'SIPARIS_TARIHI',
    'FIRMA': 'FIRMA', 'TUR': 'TUR', 'BARKOD': 'BARKOD',
    'MALIN CINSI': 'MALIN CINSI', 'ADET': 'ADET', 'FIYAT': 'FIYAT',
    'YUKLEME TARIHI': 'YUKLEME_TARIHI', 'YUKLEME_TARIHI': 'YUKLEME_TARIHI'
}

# --- CANLI DÖVİZ KURU MOTORU ---
@st.cache_data(ttl=3600)
def get_live_rates():
    rates = {"EUR_TO_USD": 1.09, "CNY_TO_USD": 0.138, "PROUNCE": "Yedek Kur Panelden Okundu"}
    try:
        response = requests.get("https://open.er-api.com/v6/latest/USD", timeout=4)
        if response.status_code == 200:
            data = response.json()
            usd_rates = data.get("rates", {})
            eur_rate = usd_rates.get("EUR")
            cny_rate = usd_rates.get("CNY")
            if eur_rate and cny_rate:
                rates["EUR_TO_USD"] = 1.0 / eur_rate
                rates["CNY_TO_USD"] = 1.0 / cny_rate
                rates["PROUNCE"] = "Canlı Kur API'den Çekildi"
    except:
        pass
    return rates

rates = get_live_rates()

# --- GELİŞMİŞ TARİH STANDARTLAŞTIRMA MOTORU ---
def strict_date_string_parser(val):
    if pd.isna(val) or val == "":
        return "BELİRTİLMEMİŞ"
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    
    val_str = str(val).strip()
    if " " in val_str:
        val_str = val_str.split()[0]
    
    val_str = val_str.replace('/', '.').replace('-', '.')
    
    for fmt in ['%Y.%m.%d', '%d.%m.%Y', '%Y.%d.%m']:
        try:
            dt = datetime.datetime.strptime(val_str, fmt)
            return dt.strftime('%Y-%m-%d')
        except:
            continue
            
    try:
        dt = pd.to_datetime(val_str, dayfirst=True, errors='coerce')
        if not pd.isna(dt):
            return dt.strftime('%Y-%m-%d')
    except:
        pass
    return "BELİRTİLMEMİŞ"

# --- VERİ TEMİZLEME VE DÖNÜŞTÜRME MOTORU ---
def clean_data(df, rates):
    df = df.loc[:, ~df.columns.duplicated()]
    for col in ['SIPARIS_TARIHI', 'YUKLEME_TARIHI']:
        if col in df.columns:
            df[col] = df[col].apply(strict_date_string_parser)
            
    available_cols = [c for c in EXPECTED_COLUMNS if c in df.columns]
    df = df[available_cols].copy()
    df = df.dropna(how='all')
    
    if 'ADET' in df.columns:
        df['ADET'] = pd.to_numeric(df['ADET'], errors='coerce').fillna(0)
        
    if 'FIYAT' in df.columns and 'FIRMA' in df.columns:
        def parse_price_details(row):
            val = row['FIYAT']
            firma_name = str(row['FIRMA']).upper().strip()
            if pd.isna(val):
                return 0.0, 0.0, '$'
            
            val_str = str(val).strip()
            currency = 'USD'
            sym_char = '$'
            
            yuan_symbols = ['¥', '￥', 'CNY', 'RMB', '元', 'CHINESE']
            euro_symbols = ['€', 'EUR', 'EURO']
            
            if 'CATHY' in firma_name or 'AECOOLY' in firma_name or any(sym in val_str for sym in yuan_symbols) or any(sym in val_str.upper() for sym in yuan_symbols):
                currency = 'CNY'
                sym_char = '¥'
            elif any(sym in val_str for sym in euro_symbols) or any(sym in val_str.upper() for sym in euro_symbols):
                currency = 'EUR'
                sym_char = '€'
                
            for clean_target in yuan_symbols + euro_symbols + ['$', 'usd', 'USD']:
                val_str = val_str.replace(clean_target, '')
            val_str = val_str.strip()
            
            if ',' in val_str and '.' in val_str:
                if val_str.find(',') > val_str.find('.'):
                    val_str = val_str.replace('.', '').replace(',', '.')
                else:
                    val_str = val_str.replace(',', '')
            elif ',' in val_str:
                val_str = val_str.replace(',', '.')
                
            try:
                numeric_price = float(val_str)
            except:
                numeric_price = 0.0
                
            if currency == 'CNY':
                usd_price = numeric_price * rates["CNY_TO_USD"]
            elif currency == 'EUR':
                usd_price = numeric_price * rates["EUR_TO_USD"]
            else:
                usd_price = numeric_price
            return usd_price, numeric_price, sym_char
            
        res = df.apply(parse_price_details, axis=1)
        df['FIYAT'] = [r[0] for r in res]
        df['ORIJINAL_FIYAT'] = [r[1] for r in res]
        df['PARA_BIRIMI'] = [r[2] for r in res]
    else:
        df['ORIJINAL_FIYAT'] = df['FIYAT'] if 'FIYAT' in df.columns else 0.0
        df['PARA_BIRIMI'] = '$'
        
    df['TOPLAM_SERMAYE'] = df['ADET'] * df['FIYAT']
    
    for text_col in ['FIRMA', 'TUR', 'MALIN CINSI', 'BARKOD', 'NAKLİYE_TÜRÜ']:
        if text_col in df.columns:
            if text_col == 'BARKOD':
                def strict_barcode_clean(x):
                    if pd.isna(x): return "BELİRTİLMEMİŞ"
                    if isinstance(x, (int, float)):
                        try:
                            if x == int(x): return str(int(x))
                            return str(x)
                        except: return str(x)
                    s = str(x).strip()
                    if s.endswith('.0'): s = s[:-2]
                    if '.' in s:
                        try:
                            f = float(s)
                            if f == int(f): return str(int(f))
                        except: pass
                    if s in ['nan', 'None', '']: return "BELİRTİLMEMİŞ"
                    return s
                df['BARKOD'] = df['BARKOD'].apply(strict_barcode_clean)
            else:
                val_series = df[text_col].fillna("BELİRTİLMEMİŞ").astype(str).str.strip()
                df[text_col] = val_series.replace({'nan': 'BELİRTİLMEMİŞ', 'None': 'BELİRTİLMEMİŞ', '': 'BELİRTİLMEMİŞ'})
    return df

# --- COKLU DOSYA VE LINK YÖNETİM MOTORU ---
@st.cache_data(ttl=600)
def get_all_data(rates):
    all_data_list = []
    pool = {tab: [] for tab in TARGET_TABS}
    
    for link in LINKS:
        try:
            response = requests.get(link, timeout=10)
            if response.status_code != 200: continue
            
            wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
            for tab in TARGET_TABS:
                if tab in wb.sheetnames:
                    sheet = wb[tab]
                    rows = list(sheet.iter_rows(values_only=False))
                    if not rows: continue
                    
                    raw_headers = [str(cell.value).strip().upper() if cell.value is not None else '' for cell in rows[0]]
                    headers = []
                    for h in raw_headers:
                        clean_h = h.replace('İ', 'I').replace('Ş', 'S').replace('Ü', 'U').replace('Ç', 'C').replace('Ğ', 'G').replace('_', ' ')
                        headers.append(HEADER_MAP.get(clean_h, h))
                        
                    try:
                        fiyat_idx = headers.index('FIYAT')
                    except ValueError:
                        fiyat_idx = -1
                        
                    data = []
                    for row in rows[1:]:
                        if all(cell.value is None for cell in row): continue
                        row_data = []
                        for idx, cell in enumerate(row):
                            if idx >= len(headers): break
                            val = cell.value
                            if idx == fiyat_idx and val is not None:
                                fmt = str(cell.number_format).upper()
                                if any(x in fmt for x in ['¥', '￥', 'CNY', '元', '804', '2052', 'E01']):
                                    val = f"¥{val}"
                                elif any(x in fmt for x in ['€', 'EUR', '40C']):
                                    val = f"€{val}"
                            row_data.append(val)
                        while len(row_data) < len(headers):
                            row_data.append(None)
                        data.append(row_data)
                        
                    df = pd.DataFrame(data, columns=headers)
                    tab_lower = tab.lower()
                    prefix = ""
                    if "has" in tab_lower: prefix = "HAS "
                    elif "meh" in tab_lower: prefix = "MEH "
                    elif "ist" in tab_lower: prefix = "IST "
                    
                    if "air" in tab_lower: df['NAKLİYE_TÜRÜ'] = prefix + "HAVA"
                    elif "sea" in tab_lower: df['NAKLİYE_TÜRÜ'] = prefix + "DENİZ"
                    else: df['NAKLİYE_TÜRÜ'] = prefix + "BELİRTİLMEMİŞ"
                    
                    df_clean = clean_data(df, rates)
                    if not df_clean.empty:
                        pool[tab].append(df_clean)
                        all_data_list.append(df_clean)
        except:
            continue
            
    full_df = pd.concat(all_data_list, ignore_index=True) if all_data_list else pd.DataFrame()
    if not full_df.empty:
        def get_clean_period(x):
            if x == "BELİRTİLMEMİŞ" or len(x) < 7: return "Bilinmeyen Dönem"
            return x[:7]
        full_df['SIPARIS_AY'] = full_df['SIPARIS_TARIHI'].apply(get_clean_period)
        
    return full_df, pool

df_dashboard, data_pool = get_all_data(rates)

# --- NAVİGASYON VE SIDEBAR YÖNETİMİ ---
st.sidebar.title("ZORE YÖNETİM PANELİ")
st.sidebar.markdown(f"**Döviz Durumu:** `{rates['PROUNCE']}`")
st.sidebar.text(f"1 EUR = {rates['EUR_TO_USD']:.4f} $")
st.sidebar.text(f"1 CNY = {rates['CNY_TO_USD']:.4f} $")
st.sidebar.markdown("---")

page = st.sidebar.radio("Sayfa Seçimi", ["1. Genel Dashboard", "2. Firma Bazlı Analiz", "3. Ham Veri"])

# --- SAYFA 1: GENEL DASHBOARD (SABİT VERİ + HAREKETLİ SİBER EKRAN) ---
if page == "1. Genel Dashboard":
    st.header("📊 Genel Dashboard - Siber İzleme Merkezi")
    
    if df_dashboard.empty:
        st.error("Veri havuzunda işlenecek kayıt bulunamadı.")
    else:
        # Üst Metrik Kartları (Orijinal yapı)
        c1, c2, c3 = st.columns(3)
        c1.metric("Toplam Sipariş Adedi (Tüm Zamanlar)", f"{int(df_dashboard['ADET'].sum()):,}")
        c2.metric("Toplam Sermaye Yatırımı (USD)", f"{df_dashboard['TOPLAM_SERMAYE'].sum():,.2f} $")
        c3.metric("Çalışılan Firma Sayısı", df_dashboard['FIRMA'].nunique())
        st.markdown("---")
        
        # VERİLER AYLARA GÖRE BÖLÜNMÜYOR, TÜM HAVUZ TEK SEFERDE HAZIRLANIYOR
        # 1. Top 10 Ürün (Adet)
        c1_df = df_dashboard.groupby('MALIN CINSI')['ADET'].sum().nlargest(10).reset_index().iloc[::-1]
        
        # 2. Top 10 Ürün ($)
        c2_df = df_dashboard.groupby('MALIN CINSI')['TOPLAM_SERMAYE'].sum().nlargest(10).reset_index().iloc[::-1]
        
        # 3. Harcama Yapılan İlk 10 Firma (Pie için)
        c3_df = df_dashboard.groupby('FIRMA')['TOPLAM_SERMAYE'].sum().nlargest(10).reset_index()
        c3_data = [{"value": round(row['TOPLAM_SERMAYE'],2), "name": row['FIRMA']} for _, row in c3_df.iterrows()]
        
        # 4. Tür Bazlı Harcama Dağılımı (Pie için)
        c4_df = df_dashboard.groupby('TUR')['TOPLAM_SERMAYE'].sum().nlargest(10).reset_index()
        c4_data = [{"value": round(row['TOPLAM_SERMAYE'],2), "name": row['TUR']} for _, row in c4_df.iterrows()]
        
        # 8. Barkod Bazlı
        df_barkod = df_dashboard[(df_dashboard['BARKOD'] != "BELİRTİLMEMİŞ") & (df_dashboard['BARKOD'].str.strip() != "")]
        c8_df = df_barkod.groupby('BARKOD').agg({'ADET': 'sum'}).nlargest(10, 'ADET').reset_index().iloc[::-1]
        
        # Çizgi (Line) Grafikleri İçin Sabit X Ekseni Hazırlığı
        valid_df = df_dashboard[df_dashboard['SIPARIS_AY'] != "Bilinmeyen Dönem"]
        all_months = sorted(valid_df['SIPARIS_AY'].unique().tolist())
        
        # 5. Aylık Firma Trendi (Sabit Seri)
        top_5_firmalar = valid_df.groupby('FIRMA')['TOPLAM_SERMAYE'].sum().nlargest(5).index
        trend_firma = valid_df[valid_df['FIRMA'].isin(top_5_firmalar)].groupby(['SIPARIS_AY', 'FIRMA'])['TOPLAM_SERMAYE'].sum().reset_index()
        c5_series = []
        for f in top_5_firmalar:
            f_data = trend_firma[trend_firma['FIRMA'] == f]
            data_arr = [f_data[f_data['SIPARIS_AY'] == m]['TOPLAM_SERMAYE'].sum() if m in f_data['SIPARIS_AY'].values else 0 for m in all_months]
            c5_series.append({"name": f, "type": "line", "smooth": True, "showSymbol": False, "data": [round(x,2) for x in data_arr]})
            
        # 6. Aylık Tür Trendi (Sabit Seri)
        top_5_turler = valid_df.groupby('TUR')['TOPLAM_SERMAYE'].sum().nlargest(5).index
        trend_tur = valid_df[valid_df['TUR'].isin(top_5_turler)].groupby(['SIPARIS_AY', 'TUR'])['TOPLAM_SERMAYE'].sum().reset_index()
        c6_series = []
        for t in top_5_turler:
            t_data = trend_tur[trend_tur['TUR'] == t]
            data_arr = [t_data[t_data['SIPARIS_AY'] == m]['TOPLAM_SERMAYE'].sum() if m in t_data['SIPARIS_AY'].values else 0 for m in all_months]
            c6_series.append({"name": t, "type": "line", "smooth": True, "showSymbol": False, "data": [round(x,2) for x in data_arr]})
            
        # 7. Aylık Toplam Sermaye (Sabit Seri)
        trend_total = valid_df.groupby('SIPARIS_AY')['TOPLAM_SERMAYE'].sum().reset_index()
        c7_data = [round(trend_total[trend_total['SIPARIS_AY'] == m]['TOPLAM_SERMAYE'].sum(), 2) if m in trend_total['SIPARIS_AY'].values else 0 for m in all_months]

        # TEK BİR JSON PAKETİ (JavaScript'e gönderilecek ve ASLA değişmeyecek)
        static_data = {
            "c1_names": c1_df['MALIN CINSI'].tolist(), "c1_vals": c1_df['ADET'].tolist(),
            "c2_names": c2_df['MALIN CINSI'].tolist(), "c2_vals": c2_df['TOPLAM_SERMAYE'].tolist(),
            "c3_data": c3_data,
            "c4_data": c4_data,
            "c5_series": c5_series,
            "c6_series": c6_series,
            "c7_months": all_months, "c7_data": c7_data,
            "c8_names": c8_df['BARKOD'].tolist(), "c8_vals": c8_df['ADET'].tolist(),
        }

        # HTML VE JAVASCRIPT TEMPLATE (SİBER EKRAN - GÖRSEL ANİMASYONLU)
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
            <style>
                body { background-color: #060913; font-family: 'Segoe UI', Tahoma, sans-serif; margin: 0; padding: 10px; overflow-x: hidden; }
                .header-box { text-align: center; border-bottom: 1px dashed rgba(0, 243, 255, 0.3); padding-bottom: 10px; margin-bottom: 20px; }
                .matrix-title { color: #ffffff; font-size: 20px; letter-spacing: 2px; margin: 0; }
                .matrix-subtitle { color: #00f3ff; font-size: 14px; margin-top: 5px; font-weight: bold; }
                .period-badge { color: #00ff66; background: rgba(0,255,102,0.15); padding: 4px 12px; border-radius: 4px; font-family: monospace; font-size: 16px; margin-left: 10px; }
                .grid-container { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
                .panel { background: rgba(4, 11, 28, 0.7); border: 1px solid rgba(0, 243, 255, 0.15); border-radius: 8px; box-shadow: 0 0 25px rgba(0, 243, 255, 0.05); height: 380px; padding: 10px; position: relative; }
                /* Parlama efekti için ufak dokunuş */
                .panel::after { content: ''; position: absolute; top:0; left:0; right:0; bottom:0; box-shadow: inset 0 0 15px rgba(0, 243, 255, 0.2); pointer-events: none; border-radius: 8px;}
            </style>
        </head>
        <body>
            <div class="header-box">
                <h2 class="matrix-title">🎬 ZORE CYBERSPACE RADAR EKRANI (SABİT VERİ + GIF ANİMASYONU)</h2>
                <div class="matrix-subtitle">
                    SİSTEM DURUMU: <span style="color: #00ff66;">TÜM VERİLER YÜKLENDİ VE SABİTLENDİ</span> | 
                    GÖRSEL MOTOR: <span class="period-badge">OTOMATİK RADAR AKTİF</span>
                </div>
            </div>
            
            <div class="grid-container">
                <div id="c1" class="panel"></div>
                <div id="c2" class="panel"></div>
                <div id="c3" class="panel"></div>
                <div id="c4" class="panel"></div>
                <div id="c5" class="panel"></div>
                <div id="c6" class="panel"></div>
                <div id="c7" class="panel"></div>
                <div id="c8" class="panel"></div>
            </div>

            <script>
                // Sabit Veri Enjeksiyonu
                const staticData = __STATIC_DATA__;
                
                const textStyle = { color: '#00f3ff', fontSize: 13, fontWeight: 'normal' };
                const axisLabelStyle = { color: '#7a92b5', fontSize: 10, width: 120, overflow: 'truncate' };
                const splitLineStyle = { lineStyle: { color: 'rgba(0,243,255,0.05)' } };

                const charts = {
                    c1: echarts.init(document.getElementById('c1')), c2: echarts.init(document.getElementById('c2')),
                    c3: echarts.init(document.getElementById('c3')), c4: echarts.init(document.getElementById('c4')),
                    c5: echarts.init(document.getElementById('c5')), c6: echarts.init(document.getElementById('c6')),
                    c7: echarts.init(document.getElementById('c7')), c8: echarts.init(document.getElementById('c8'))
                };

                const gradBlue = new echarts.graphic.LinearGradient(0,0,1,0, [{offset:0, color:'#0033ff'}, {offset:1, color:'#00f3ff'}]);
                const gradOrange = new echarts.graphic.LinearGradient(0,0,1,0, [{offset:0, color:'#ff5500'}, {offset:1, color:'#ffaa00'}]);
                const gradPink = new echarts.graphic.LinearGradient(0,0,1,0, [{offset:0, color:'#9900ff'}, {offset:1, color:'#ff00ff'}]);

                // 1. Sabit Verilerin Grafiklere Yerleştirilmesi (Veri asla değişmez)
                charts.c1.setOption({ title: { text: '1. En Çok Sipariş Edilen 10 Ürün (Adet)', textStyle: textStyle }, tooltip: { trigger: 'axis' }, grid: { left: '25%', right: '5%', bottom: '5%', top: '15%' }, xAxis: { type: 'value', splitLine: splitLineStyle, axisLabel: {color:'#7a92b5'} }, yAxis: { type: 'category', data: staticData.c1_names, axisLabel: axisLabelStyle }, series: [{ type: 'bar', data: staticData.c1_vals, itemStyle: { color: gradBlue, borderRadius: [0,4,4,0] } }] });
                charts.c2.setOption({ title: { text: '2. En Çok Sermaye Yatırılan 10 Ürün ($)', textStyle: textStyle }, tooltip: { trigger: 'axis' }, grid: { left: '25%', right: '5%', bottom: '5%', top: '15%' }, xAxis: { type: 'value', splitLine: splitLineStyle, axisLabel: {color:'#7a92b5'} }, yAxis: { type: 'category', data: staticData.c2_names, axisLabel: axisLabelStyle }, series: [{ type: 'bar', data: staticData.c2_vals, itemStyle: { color: gradOrange, borderRadius: [0,4,4,0] } }] });
                
                // Pastalara 'animationDurationUpdate: 0' ekliyoruz ki radar gibi pürüzsüz dönsünler
                charts.c3.setOption({ title: { text: '3. Harcama Yapılan İlk 10 Firma (USD)', textStyle: textStyle }, tooltip: { trigger: 'item' }, series: [{ type: 'pie', radius: [20, 100], center: ['50%', '55%'], roseType: 'area', itemStyle: { borderRadius: 4 }, label: { color: '#7a92b5', fontSize: 10 }, data: staticData.c3_data, animationDurationUpdate: 0 }] });
                charts.c4.setOption({ title: { text: '4. Tür Bazlı Harcama Dağılımı (USD)', textStyle: textStyle }, tooltip: { trigger: 'item' }, series: [{ type: 'pie', radius: ['40%', '70%'], center: ['50%', '55%'], itemStyle: { borderRadius: 5, borderColor: '#060913', borderWidth: 2 }, label: { color: '#7a92b5', fontSize: 11 }, data: staticData.c4_data, animationDurationUpdate: 0 }] });
                
                charts.c5.setOption({ title: { text: '5. Aylık Firma Harcama Trendi (Top 5)', textStyle: textStyle }, tooltip: { trigger: 'axis' }, grid: { left: '10%', right: '5%', bottom: '10%', top: '20%' }, xAxis: { type: 'category', data: staticData.c7_months, axisLabel: {color:'#7a92b5'} }, yAxis: { type: 'value', splitLine: splitLineStyle, axisLabel: {color:'#7a92b5'} }, series: staticData.c5_series });
                charts.c6.setOption({ title: { text: '6. Aylık Tür Harcama Trendi (Top 5)', textStyle: textStyle }, tooltip: { trigger: 'axis' }, grid: { left: '10%', right: '5%', bottom: '10%', top: '20%' }, xAxis: { type: 'category', data: staticData.c7_months, axisLabel: {color:'#7a92b5'} }, yAxis: { type: 'value', splitLine: splitLineStyle, axisLabel: {color:'#7a92b5'} }, series: staticData.c6_series });
                charts.c7.setOption({ title: { text: '7. Aylık Toplam Sermaye Akışı ($)', textStyle: textStyle }, tooltip: { trigger: 'axis' }, grid: { left: '10%', right: '5%', bottom: '10%', top: '15%' }, xAxis: { type: 'category', data: staticData.c7_months, axisLabel: {color:'#7a92b5'}, splitLine: splitLineStyle }, yAxis: { type: 'value', splitLine: splitLineStyle, axisLabel: {color:'#7a92b5'} }, series: [{ type: 'line', smooth: true, areaStyle: { color: 'rgba(0, 243, 255, 0.2)' }, lineStyle: { color: '#00f3ff', width: 3, shadowBlur: 10, shadowColor: 'rgba(0,243,255,0.8)' }, itemStyle: { color: '#00f3ff' }, data: staticData.c7_data }] });
                charts.c8.setOption({ title: { text: '8. Barkod Bazlı Top 10 Ürün (Adet)', textStyle: textStyle }, tooltip: { trigger: 'axis' }, grid: { left: '20%', right: '5%', bottom: '5%', top: '15%' }, xAxis: { type: 'value', splitLine: splitLineStyle, axisLabel: {color:'#7a92b5'} }, yAxis: { type: 'category', data: staticData.c8_names, axisLabel: axisLabelStyle }, series: [{ type: 'bar', data: staticData.c8_vals, itemStyle: { color: gradPink, borderRadius: [0,4,4,0] } }] });

                // 2. GIF BENZERİ GÖRSEL ANİMASYON MOTORU (VERİ DEĞİŞTİRMEZ, SADECE GÖRSELLİK)
                
                // A) Pastalar için kesintisiz dönüş (Radar Efekti)
                let currentAngle = 0;
                setInterval(() => {
                    currentAngle = (currentAngle - 1) % 360; // Geriye doğru pürüzsüz dönüş
                    charts.c3.setOption({ series: [{ startAngle: currentAngle }] });
                    charts.c4.setOption({ series: [{ startAngle: currentAngle }] });
                }, 50); // Çok hızlı tetiklenerek dönme hissi verir

                // B) Barlar ve Çizgiler İçin Otomatik Tarayıcı (Scanner Efekti)
                let scanIndex = 0;
                setInterval(() => {
                    let totalBars = staticData.c1_names.length;
                    let totalMonths = staticData.c7_months.length;
                    
                    // Bar grafiklerinde yukarıdan aşağıya vurgulama (Highlighting)
                    if (totalBars > 0) {
                        let bIdx = scanIndex % totalBars;
                        [charts.c1, charts.c2, charts.c8].forEach(chart => {
                            chart.dispatchAction({ type: 'downplay' });
                            chart.dispatchAction({ type: 'highlight', seriesIndex: 0, dataIndex: bIdx });
                            chart.dispatchAction({ type: 'showTip', seriesIndex: 0, dataIndex: bIdx });
                        });
                    }

                    // Çizgi grafiklerinde soldan sağa zaman taraması (Heartbeat Efekti)
                    if (totalMonths > 0) {
                        let mIdx = scanIndex % totalMonths;
                        [charts.c5, charts.c6, charts.c7].forEach(chart => {
                            chart.dispatchAction({ type: 'showTip', seriesIndex: 0, dataIndex: mIdx });
                        });
                    }

                    scanIndex++;
                }, 1500); // Her 1.5 saniyede bir sonrakine geçer

                // Ekran yeniden boyutlandırma
                window.addEventListener('resize', () => {
                    Object.values(charts).forEach(c => c.resize());
                });
            </script>
        </body>
        </html>
        """
        
        html_ready = html_template.replace("__STATIC_DATA__", json.dumps(static_data))
        
        # Matrix arayüzünü Streamlit'e ekleme
        st.components.v1.html(html_ready, height=1700, scrolling=False)


# --- SAYFA 2: FİRMA BAZLI ANALİZ ---
elif page == "2. Firma Bazlı Analiz":
    st.header("🏢 Firma Bazlı Analiz")
    if df_dashboard.empty:
        st.error("Veri havuzu boş.")
    else:
        firmalar = sorted([str(f) for f in df_dashboard['FIRMA'].unique() if str(f) != "BELİRTİLMEMİŞ"])
        if not firmalar:
            st.warning("Analiz edilecek geçerli bir firma kaydı bulunamadı.")
        else:
            selected_firma = st.selectbox("Analiz edilecek firmayı seçin", firmalar)
            firma_df = df_dashboard[df_dashboard['FIRMA'] == selected_firma]
            
            c1, c2, c3 = st.columns(3)
            c1.metric(f"{selected_firma} Toplam Alım (Adet)", f"{int(firma_df['ADET'].sum()):,}")
            c2.metric(f"{selected_firma} Toplam Ciro (USD)", f"{firma_df['TOPLAM_SERMAYE'].sum():,.2f} $")
            tur_counts = firma_df.groupby('TUR')['ADET'].sum()
            en_cok_tur = tur_counts.idxmax() if not tur_counts.empty and tur_counts.sum() > 0 else "Veri Yok"
            c3.metric("En Çok Aldığı Tür", en_cok_tur)
            
            import plotly.express as px
            col_a, col_b = st.columns(2)
            if not firma_df.empty and firma_df['TOPLAM_SERMAYE'].sum() > 0:
                kategori_ozet = firma_df.groupby('TUR')['TOPLAM_SERMAYE'].sum().reset_index()
                if len(kategori_ozet) > 6:
                    en_buyuk_6 = kategori_ozet.nlargest(6, 'TOPLAM_SERMAYE')['TUR'].tolist()
                    firma_df_pie = firma_df.copy()
                    firma_df_pie['TUR_GRAFIK'] = firma_df_pie['TUR'].apply(lambda x: x if x in en_buyuk_6 else 'DİĞER')
                else:
                    firma_df_pie = firma_df.copy()
                    firma_df_pie['TUR_GRAFIK'] = firma_df_pie['TUR']
                
                fig_a = px.pie(firma_df_pie, values='TOPLAM_SERMAYE', names='TUR_GRAFIK', title=f"{selected_firma} Ürün Kategorisi Dağılımı (İlk 6 + Diğer)", hole=0.4)
                fig_a.update_traces(textinfo='label+percent')
                col_a.plotly_chart(fig_a, use_container_width=True)
            else:
                col_a.info("Grafik için yeterli veri yok.")
                
            trend_data_all = firma_df.groupby('SIPARIS_AY')['TOPLAM_SERMAYE'].sum().reset_index().sort_values('SIPARIS_AY')
            if not trend_data_all.empty and trend_data_all['TOPLAM_SERMAYE'].sum() > 0:
                fig_b = px.bar(trend_data_all, x='SIPARIS_AY', y='TOPLAM_SERMAYE', title=f"{selected_firma} Dönemsel Alım Trendi ($)", color='TOPLAM_SERMAYE')
                fig_b.update_layout(xaxis_type='category')
                col_b.plotly_chart(fig_b, use_container_width=True)
            else:
                col_b.info("Zaman trendi grafik verisi bulunamadı.")
                
            st.markdown("---")
            st.subheader(f"🔍 {selected_firma} Sipariş Listesinde Barkod Sorgulama")
            search_barcode = st.text_input("Barkod Yazın (Varmı / Yokmu Kontrolü):", placeholder="Kontrol etmek istediğiniz barkodu buraya girin...").strip()
            display_df = firma_df.copy()
            
            if search_barcode:
                search_res = display_df[display_df['BARKOD'].str.contains(search_barcode, case=False, na=False)]
                if not search_res.empty:
                    st.success(f"✅ Barkod Bulundu! Bu firmaya ait listede aradığınız barkod ile eşleşen {len(search_res)} adet kayıt var.")
                    display_df = search_res
                else:
                    st.error("❌ Barkod Bulunamadı! Bu firmanın ham veri listesinde yazdığınız barkod mevcut değil.")
                    
            st.markdown(f"**{selected_firma} Veri Listesi:**")
            display_df_formatted = display_df.copy()
            display_df_formatted['FIYAT'] = display_df_formatted['FIYAT'].map('{:,.2f} $'.format)
            display_df_formatted['TOPLAM_SERMAYE'] = display_df_formatted['TOPLAM_SERMAYE'].map('{:,.2f} $'.format)
            
            drop_cols = [c for c in ['ORIJINAL_FIYAT', 'PARA_BIRIMI'] if c in display_df_formatted.columns]
            if drop_cols:
                display_df_formatted = display_df_formatted.drop(columns=drop_cols)
            st.dataframe(display_df_formatted.sort_values(by='SIPARIS_TARIHI', ascending=False), use_container_width=True, hide_index=True)

# --- SAYFA 3: HAM VERİ ---
elif page == "3. Ham Veri":
    st.header("📋 Ham Veri Havuzu")
    tabs = st.tabs(TARGET_TABS)
    for i, tab_ui in enumerate(tabs):
        with tab_ui:
            tab_name = TARGET_TABS[i]
            df_list = data_pool[tab_name]
            if df_list:
                combined_df = pd.concat(df_list, ignore_index=True).drop_duplicates()
                raw_display = combined_df.copy()
                raw_display['FIYAT'] = raw_display['FIYAT'].map('{:,.2f} $'.format)
                raw_display['TOPLAM_SERMAYE'] = raw_display['TOPLAM_SERMAYE'].map('{:,.2f} $'.format)
                drop_cols = [c for c in ['ORIJINAL_FIYAT', 'PARA_BIRIMI'] if c in raw_display.columns]
                if drop_cols:
                    raw_display = raw_display.drop(columns=drop_cols)
                st.dataframe(raw_display, use_container_width=True, hide_index=True)
            else:
                st.warning(f"Bu sekme ({tab_name}) için veri bulunamadı.")